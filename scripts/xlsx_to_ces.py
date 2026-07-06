#!/usr/bin/env python3
"""Convert the Chart Review Agent's chart_reviews.xlsx into CES bot batches.

Reads the "Chart Reviews" sheet the Ninth Brain Chart Review Agent writes and
emits the JSON the CES QA Review Queue imports (docs/bot-bridge.md):

    python3 xlsx_to_ces.py chart_reviews.xlsx
    python3 xlsx_to_ces.py chart_reviews.xlsx --dry-run

Because CES imports into one operation's period at a time, reviews are split
by Business Unit — e.g. ces_batch_kc.json and ces_batch_linn.json — unless the
report only contains one unit (then it's just ces_batch.json). Import each
file into the matching operation's period: QA -> period -> Add charts -> Bot
reviews.

What gets carried over per review:
  - Run ID -> incident number (the match key), Run Date -> date,
    Care Level -> acuity
  - Q1–Q15 answers -> the CES rubric criteria q1–q15. Q14/Q15 are asked in
    reverse on the Ninth Brain form, so their answers are inverted (a "No"
    on "Were there any near misses…" imports as Met).
  - CES computes the weighted score from the criteria.
  - flagged = true when Q14 or Q15 was answered "Yes" on the form
    (safety concern / needs leadership review).
  - Synopsis, Findings, matched rules, any category follow-ups (STEMI /
    Trauma / Stroke / AMS), and rationales for No answers -> notes.

The "Code 3 Case Presentations" sheet is a different work product (CQMP
ground-vs-air audit) and is not part of the QA queue; it is skipped.

Files without the bot's columns fall back to a generic header auto-detect
(incident/score/provider/notes columns), same aliases as the CES importer.

Requires: openpyxl (already installed for the bot).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from typing import Any, Optional

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required (pip3 install openpyxl)")

# Ninth Brain form questions asked in the negative — invert on import.
INVERTED_QUESTIONS = {"Q14", "Q15"}
NUM_QUESTIONS = 15

# Business Unit -> CES operation id.
UNIT_PATTERNS = [
    ("kc", re.compile(r"kansas\s*city", re.I)),
    ("linn", re.compile(r"linn", re.I)),
    ("cass", re.compile(r"cass", re.I)),
]


def norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def iso_date(v: str) -> str:
    """Normalize '6/11/2026' (and already-ISO strings) to yyyy-mm-dd."""
    v = v.strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", v)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return v


def split_answer(cell: str) -> tuple[str, str]:
    """'Yes — rationale…' -> ('Yes', 'rationale…'); tolerate -, – and em-dash."""
    m = re.match(r"^\s*(Yes|No)\s*[—–-]\s*(.*)$", cell, re.S)
    if m:
        return m.group(1), m.group(2).strip()
    stripped = cell.strip()
    if stripped.lower() in ("yes", "no"):
        return stripped.capitalize(), ""
    return "", stripped


def unit_of(business_unit: str) -> str:
    for op, pat in UNIT_PATTERNS:
        if pat.search(business_unit):
            return op
    return "other"


def read_sheet(ws) -> tuple[list[str], list[list[str]]]:
    rows = [[cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(v for v in r)]
    if not rows:
        return [], []
    return rows[0], rows[1:]


# ── Bot-report parser ────────────────────────────────────────────────────────

def parse_bot_report(headers: list[str], data: list[list[str]]) -> list[dict[str, Any]]:
    col = {h: i for i, h in enumerate(headers)}

    def get(row: list[str], header: str) -> str:
        i = col.get(header)
        return row[i] if i is not None and i < len(row) else ""

    q_headers = [f"Q{i}" for i in range(1, NUM_QUESTIONS + 1) if f"Q{i}" in col]
    core = {"Reviewed At", "Run ID", "Run Date", "Business Unit", "Matched Rules",
            "Review Types", "Care Level", "Synopsis", "Findings", *q_headers}
    followup_headers = [h for h in headers if h and h not in core]

    reviews: list[dict[str, Any]] = []
    for row in data:
        inc = get(row, "Run ID")
        if not inc:
            continue

        criteria: dict[str, str] = {}
        no_rationales: list[str] = []
        flagged = False
        for qh in q_headers:
            answer, rationale = split_answer(get(row, qh))
            if not answer:
                continue
            good = answer == "Yes"
            if qh in INVERTED_QUESTIONS:
                if answer == "Yes":
                    flagged = True  # safety concern / needs leadership review
                good = not good
            criteria[qh.lower()] = "met" if good else "not_met"
            if not good and rationale:
                no_rationales.append(f"{qh}: {rationale}")

        note_parts: list[str] = []
        synopsis = get(row, "Synopsis")
        findings = get(row, "Findings")
        rules = get(row, "Matched Rules")
        if synopsis:
            note_parts.append(synopsis)
        if findings:
            note_parts.append(f"Findings: {findings}")
        if rules:
            note_parts.append(f"Matched rules: {rules}")
        if no_rationales:
            note_parts.append("Deficiencies:\n" + "\n".join(no_rationales))
        followups = [f"{h}: {get(row, h)}" for h in followup_headers if get(row, h)]
        if followups:
            note_parts.append("Category review:\n" + "\n".join(followups))

        review: dict[str, Any] = {
            "incidentNumber": inc,
            "reviewer": "Chart Review Agent",
            "_unit": unit_of(get(row, "Business Unit")),
        }
        d = get(row, "Run Date")
        if d:
            review["date"] = iso_date(d)
        acuity = get(row, "Care Level")
        if acuity:
            review["acuity"] = acuity
        if criteria:
            review["criteria"] = criteria
        review["flagged"] = flagged
        if note_parts:
            review["notes"] = "\n\n".join(note_parts)
        reviews.append(review)
    return reviews


# ── Generic fallback parser (non-bot spreadsheets) ──────────────────────────

FIELD_ALIASES: dict[str, list[str]] = {
    "incidentNumber": ["incidentnumber", "incident", "run", "runnumber", "runid", "pcr", "report", "callnumber", "call", "id"],
    "date": ["date", "calldate", "servicedate", "rundate", "dispatchdate"],
    "provider": ["provider", "primary", "medic", "clinician", "author", "crewlead", "attendant"],
    "crew": ["crew", "unit", "partner", "vehicle"],
    "chiefComplaint": ["chiefcomplaint", "complaint", "chief", "impression", "reason", "nature"],
    "acuity": ["acuity", "priority", "severity", "carelevel", "level"],
    "scorePct": ["scorepct", "score", "qascore", "percent", "scorepercent", "overall", "grade"],
    "flagged": ["flagged", "flag", "followup", "coaching"],
    "notes": ["notes", "summary", "findings", "synopsis", "comment", "comments", "narrative", "feedback"],
    "reviewer": ["reviewer", "agent", "model"],
}

TRUTHY = {"true", "yes", "1", "y", "flag", "flagged"}


def parse_generic(headers: list[str], data: list[list[str]]) -> list[dict[str, Any]]:
    normed = [norm(h) for h in headers]
    mapping: dict[str, int] = {}
    used: set[int] = set()
    for field, aliases in FIELD_ALIASES.items():
        for i, h in enumerate(normed):
            if i not in used and h in aliases:
                mapping[field] = i
                used.add(i)
                break
        else:
            for i, h in enumerate(normed):
                if i in used or not h:
                    continue
                if any(a in h for a in aliases):
                    mapping[field] = i
                    used.add(i)
                    break
    if "incidentNumber" not in mapping:
        sys.exit(
            "Could not find an incident / run number column.\n"
            f"Headers seen: {headers}"
        )
    reviews: list[dict[str, Any]] = []
    for row in data:
        get = lambda f: row[mapping[f]] if f in mapping and mapping[f] < len(row) else ""
        inc = get("incidentNumber")
        if not inc:
            continue
        review: dict[str, Any] = {"incidentNumber": inc, "_unit": "all"}
        for f in ("date", "provider", "crew", "chiefComplaint", "acuity", "notes", "reviewer"):
            v = get(f)
            if v:
                review[f] = iso_date(v) if f == "date" else v
        score = get("scorePct")
        if score:
            try:
                n = float(str(score).rstrip("%"))
                review["scorePct"] = round(n * 100) if n <= 1 else round(n)
            except ValueError:
                pass
        flag = get("flagged")
        if flag:
            review["flagged"] = norm(flag) in TRUTHY
        reviews.append(review)
    return reviews


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="Convert the Chart Review Agent's .xlsx into CES bot batches.")
    ap.add_argument("xlsx", help="Path to chart_reviews.xlsx")
    ap.add_argument("-o", "--out", default="ces_batch", help="Output basename (default ces_batch)")
    ap.add_argument("--sheet", help="Worksheet name (default: auto)")
    ap.add_argument("--unit", choices=["kc", "linn", "cass"], help="Only export this operation's reviews")
    ap.add_argument("--dry-run", action="store_true", help="Report what would be written, write nothing")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True, read_only=True)
    if args.sheet:
        ws = wb[args.sheet]
    elif "Chart Reviews" in wb.sheetnames:
        ws = wb["Chart Reviews"]
    else:
        ws = wb.worksheets[0]
    headers, data = read_sheet(ws)
    wb.close()
    if not data:
        sys.exit(f"No data rows found in sheet '{ws.title}'.")

    is_bot_report = "Run ID" in headers and "Q1" in headers
    if is_bot_report:
        print(f"Detected Chart Review Agent report ({len(data)} rows).")
        reviews = parse_bot_report(headers, data)
    else:
        print("Not a Chart Review Agent report — using generic column detection.")
        reviews = parse_generic(headers, data)

    # Group by operation.
    by_unit: dict[str, list[dict[str, Any]]] = {}
    for r in reviews:
        unit = r.pop("_unit", "all")
        by_unit.setdefault(unit, []).append(r)

    if args.unit:
        by_unit = {k: v for k, v in by_unit.items() if k == args.unit}
        if not by_unit:
            sys.exit(f"No reviews found for unit '{args.unit}'.")

    single = len(by_unit) == 1
    for unit, unit_reviews in sorted(by_unit.items()):
        flagged = sum(1 for r in unit_reviews if r.get("flagged"))
        out = f"{args.out}.json" if single else f"{args.out}_{unit}.json"
        print(f"  {unit}: {len(unit_reviews)} reviews ({flagged} flagged) -> {out}")
        if args.dry_run:
            continue
        with open(out, "w", encoding="utf-8") as fh:
            json.dump(
                {"source": "ninth-brain-chart-review-agent", "version": 1, "reviews": unit_reviews},
                fh,
                indent=2,
            )
    if args.dry_run:
        print("Dry run — nothing written.")
    else:
        print("Import each file into the matching operation's period: QA -> period -> Add charts -> Bot reviews.")


if __name__ == "__main__":
    main()
